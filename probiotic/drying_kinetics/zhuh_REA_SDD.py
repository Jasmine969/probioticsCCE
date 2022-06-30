import numpy as np
import pandas as pd
from scipy.integrate import ode
from matplotlib import pyplot as plt
from sklearn.metrics import r2_score
from scipy.interpolate import interp1d
import openpyxl as opx
from string import ascii_uppercase as upper

Mair = 28.85
Mvapor = 18.0153
P = 101325.
RR = 8314.
pi = np.pi
Y = 0.0001
Tp0 = 293.15
R = 8.314
WaterCpCoeffs = [15341.1046350264, -116.019983347211, 0.451013044684985,
                 -0.000783569247849015, 5.20127671384957e-07]
WateRhoCoeffs = [98.343885, 0.30542, 647.13, 0.081]
HLCoeffs = [647.13, 2889425.47876769, 0.3199, -0.212, 0.25795, 0.]
AirMuCoeffs = [1.5061e-06, 6.16e-08, -1.819e-11, 0., 0.]
AirCpCoeffs = [948.76, 0.39171, -0.00095999, 1.393e-06, -6.2029e-10]
AirRhoCoeffs = [4.0097, -0.016954, 3.3057e-05, -3.0042e-08, 1.0286e-11]
AirKappaCoeffs = [0.0025219, 8.506e-05, -1.312e-08, 0., 0.]
VaporMuCoeffs = [1.5068e-06, 6.1598e-08, -1.8188e-11, 0, 0]
VaporRhoCoeffs = [2.5039, -0.010587, 2.0643e-05, -1.8761e-08, 6.4237e-12]
VaporKappaCoeffs = [0.0037972, 0.00015336, -1.1859e-08, 0, 0]
VaporCpCoeffs = [1563.1, 1.604, -0.0029334, 3.2168e-06, -1.1571e-09]

VarList = lambda x: [1., x, x ** 2, x ** 3, x ** 4]
shrink_itp = interp1d([0.1, 0.2, 0.3], [0.51, 0.59, 0.69])


def calcWaterRho(T):
    """
    计算温度T(K)下水的密度(kg/m3)
    """
    a_, b_, c_, d_ = WateRhoCoeffs
    return a_ / pow(b_, 1 + pow(1 - T / c_, d_))


def calcMilkRho():
    """
    计算温度T下牛奶的密度
    """
    return 1435.


def calcWaterHL(T):
    """
    计算温度T下的水的蒸发焓(J/kg)
    """
    Tc_, a_, b_, c_, d_, e_ = HLCoeffs
    Tr = T / Tc_
    return a_ * pow(1 - Tr, ((e_ * Tr + d_) * Tr + c_) * Tr + b_)


def calcDropletCp(T, py):
    """
    计算温度T下固含量为py的脱脂奶的比热容(J/(kg·K))
    """
    T = T - 273.15
    CpLactose = 1548.84 + 1.9625 * T - 0.0059399 * T ** 2.
    CpProtein = 2008.2 + 1.2089 * T - 0.0013129 * T ** 2.
    CpFat = 1984.2 + 1.4733 * T - 0.0048008 * T ** 2.
    CpMineral = 1092.6 + 1.8896 * T - 0.0036817 * T ** 2.
    cpMilk = CpLactose * 0.498 + CpProtein * 0.365 + CpFat * 0.006 + CpMineral * 0.093  # + CpWater*0.038
    # Cpd = py * CpWater + (1 - py) * cpMilk
    # print(Cpd)
    return cpMilk


def calcPhysicalProperities(Tf, Y):
    """
    计算膜温度Tf(K)下水蒸气质量分数为Y的湿空气的密度、粘度(Pa·s)、比热容、热传导系数W/(m·K)
    """
    TfList = [1., Tf, Tf ** 2, Tf ** 3, Tf ** 4]
    Mmixture = 1. / (Y / Mvapor + (1 - Y) / Mair)
    rho = P * Mmixture / (RR * Tf)  # 密度
    mu = (1 - Y) * np.dot(AirMuCoeffs, TfList) + Y * np.dot(VaporMuCoeffs, TfList)  # 黏度
    kappa = (1 - Y) * np.dot(AirKappaCoeffs, TfList) + Y * np.dot(VaporKappaCoeffs, TfList)  # 热传导系数
    cp = ((1 - Y) * np.dot(AirCpCoeffs, TfList) + Y * np.dot(VaporCpCoeffs, TfList))  # cp

    return [rho, mu, cp, kappa]


def sdd(l, w, Ta, pMass0, dp0, va, ms, ws, Y):
    """
    计算单液滴脱脂奶干燥主程序
    """
    # 活化能特征方程要用
    a = 0.998
    b = 1.405
    d = 0.930

    Mo = 0.06156
    Co = 0.001645  # GAB
    Hi = 24831.
    Ko = 5.71  # GAB
    Hii = -5118
    TpNow, pMass = w
    Mw = [Mair, Mvapor]
    Ts = (2. * TpNow + Ta) / 3.
    px = (pMass - ms) / ms  # 液滴、颗粒的干基含水率
    py = (pMass - ms) / pMass  # 液滴、颗粒中水分的质量分数
    diameter = 0
    if 0.05 <= ws <= 0.15:
        deltaMpByRhoWater = (pMass0 - pMass) / calcWaterRho(TpNow)
        diameter = ((np.pi / 6. * dp0 ** 3. - deltaMpByRhoWater) * 6. / np.pi) ** (1. / 3.)
    elif ws <= 0.3:
        alpha = shrink_itp(ws)
        diameter = dp0 * (alpha + (1 - alpha) * (px / ((1 - ws) / ws)))
    else:
        print("error \n")

    Ap = pi * diameter ** 2.

    fPsat = np.exp(23.365 - 3919.9 / (Ta - 42.062))  # 算RH用
    Psat = np.exp(23.365 - 3919.9 / (TpNow - 42.062))  # 算RH
    AH = Y / (1 - Y)  # 绝对湿度
    RH = (P * AH) / (0.622 * fPsat + AH * fPsat)  # 相对湿度
    # ----------------------------------------
    C = Co * np.exp(Hi / (8.314 * Ta))
    K = Ko * np.exp(Hii / (8.314 * Ta))
    pxe = (C * K * Mo * RH) / ((1 - K * RH) * (1 - K * RH + C * K * RH))
    lnRH = np.log(RH)
    deltaEve = -(Ta * 8314. * lnRH)  # 平衡活化能

    Yavg = Y
    nx = px - pxe
    rhoS = None
    if nx >= 0:
        deltaEv = 0
        if 0.05 <= ws <= 0.25:
            deltaEv = deltaEve * a * np.exp(-b * nx ** d)
        elif 0.25 <= ws <= 0.3:
            deltaEv = deltaEve * (3.0318e-2 * nx ** 4 - 0.26637 * nx ** 3
                                  + 0.85762 * nx ** 2 - 1.3635 * nx + 0.99609)

        Pvs = Psat * np.exp(-deltaEv / (8314. * TpNow))  # 先算表面的压力，再在下行换算成密度
        rhoS = Pvs * Mvapor / 8314. / TpNow  # rho = pM/RT
        Yw = (Pvs / P * Mvapor) / (Pvs / P * Mair + (1 - Pvs / P) * Mair)
        Yavg = (Yw + Y) / 2.

    rho, mu, cp, kappa = calcPhysicalProperities(Ts, Yavg)
    Xvapor = Y / Mw[1] / (Y / Mw[1] + (1 - Y) / Mw[0])
    Dab = 3.564e-10 * ((Ts * 2.0) ** 1.75)
    Cpd = calcDropletCp(TpNow, py)
    HL = calcWaterHL(TpNow)

    Re = diameter * va * rho / mu
    Sc = mu / rho / Dab
    Pr = mu * cp / kappa
    Nu = 2.04 + 0.62 * Re ** 0.5 * Pr ** (1. / 3.)
    Sh = 1.63 + 0.54 * Re ** 0.5 * Sc ** (1. / 3.)

    hm = Sh * Dab / diameter
    h = Nu * kappa / diameter
    dmdt = 0.0
    rhoF = Xvapor * P * Mvapor / (8314. * Ta)  # rho_{v,b}
    if nx >= 0:
        dmdt = hm * Ap * (rhoS - rhoF)
    dTpdt = (h * (Ta - TpNow) * Ap - (dmdt * HL)) / (pMass * Cpd)
    # -----------------------------------
    # C1 = 1.63 + 0.54 * (va * rho / mu) ** 0.5 * Sc ** (1 / 3)
    # print(1)
    # wb = opx.load_workbook('E:\\probioticsCCE\\probiotic\\vary_inp_test\\vd1.xlsx')
    # ws = wb['Sheet3']
    # write_data = [C1, diameter, hm, -dmdt, Ap, rhoS, rhoF, rhoS-rhoF]
    # row = ws.max_row + 1
    # for ind, wr in enumerate(write_data):
    #     ws[upper[ind] + str(row)] = wr
    # wb.save('vd1.xlsx')
    # print(row)

    return [dTpdt, -dmdt]


def read_temp_x(groups: list):
    data = []
    for i in groups:
        # datum = pd.read_excel('../excel/raw_1s.xlsx', sheet_name='Sheet' + str(i))
        # data.append(datum.to_numpy()[:, :3])
        datum = pd.read_excel('../excel/itp_ft_s.xlsx', sheet_name='Sheet' + str(i))
        data.append(datum.to_numpy()[:, [0, 5]])
    return data


def calcDropRho(Tp, ws):
    rhoDrop = (1 - ws) * calcWaterRho(Tp) + ws * calcMilkRho()  # 颗粒密度
    return rhoDrop


def gen_sdd_data(Ta, va, ws, vd, dur):
    """
    计算任意环境温度、风速、RSM含量、液滴体积、干燥持续时间的Td和X数据
    """
    # wb = opx.load_workbook('E:\\python_code\\probioticsCCE\\probiotic\\vary_inp_test\\vd1.xlsx')
    # wb = opx.Workbook()
    # wb.create_sheet('Sheet3')
    # wb.save('vd1.xlsx')

    rhoDrop0 = calcDropRho(Tp0, ws)
    dp0 = (6 * vd / pi) ** (1 / 3)
    pMass0 = dp0 ** 3 * pi / 6. * rhoDrop0  # 颗粒质量
    ms = pMass0 * ws  # 固形物质量
    w0 = [Tp0, pMass0]
    time = [0.]
    result = [w0]
    r0 = ode(sdd).set_integrator('lsoda', method='bdf')
    #  Y是水蒸气质量分数
    r0.set_initial_value(w0, 0).set_f_params(Ta, pMass0, dp0, va, ms, ws, Y)
    dur = float(dur)
    while r0.successful() and r0.t < dur:
        r0.integrate(r0.t + 1)
        time.append(r0.t)
        result.append(r0.y)
    result = np.array(result)
    a, b = result.shape
    if a < dur - 6:
        return np.nan
    time = np.array(time)
    temp = result[:, 0]
    mois = result[:, 1] / ms - 1
    return np.c_[time, temp, mois]
